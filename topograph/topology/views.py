import json

from django.http import HttpResponse
from django.urls import reverse_lazy
from django.shortcuts import render, get_object_or_404, redirect, HttpResponseRedirect, reverse
from django.views.generic import DetailView, ListView, UpdateView, CreateView, DeleteView
from django.contrib.auth.views import LoginView
from django.contrib.auth import logout, login
from django.contrib.auth.models import User

from .models import *
from .forms import *
from .parsers import get_lsdb
import datetime
from .tasks import calc_layout
from openpyxl import Workbook



# Create your views here.


def index(request):
    return render(request, 'topology/home.html')

def logout_user(request):
    logout(request)
    return redirect('login')


def topology_export(request, pk):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(
        date=datetime.datetime.now().strftime('%Y-%m-%d'),
    )

    nodes_queryset = Node.objects.select_related().filter(topology__pk__exact=pk)
    edges_queryset = Edge.objects.select_related().filter(begin__topology=pk)

    book = Workbook()
    sheet_nodes = book.active
    sheet_nodes.title = "nodes"
    sheet_edges = book.create_sheet("edges")

    # Define the titles for columns
    node_columns = [
        'id',
        'REMOTE_HOST',
        'id',
        'REMOTE_HOST',
        'os',
        'version',
        'x',
        'y'
    ]
    edge_columns = [
        'node1',
        'node2',
        'cost',
        'meta',
    ]

    row_num = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(node_columns, 1):
        cell = sheet_nodes.cell(row=row_num, column=col_num)
        cell.value = column_title

    for col_num, column_title in enumerate(edge_columns, 1):
        cell = sheet_edges.cell(row=row_num, column=col_num)
        cell.value = column_title

    for node in nodes_queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            node.pk,
            node.label,
            node.pk,
            node.label,
            'os:',
            'version:',
            node.x,
            node.y,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = sheet_nodes.cell(row=row_num, column=col_num)
            cell.value = cell_value

    row_num = 1
    for edge in edges_queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            edge.begin.pk,
            edge.end.pk,
            edge.cost,
            edge.meta_data,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = sheet_edges.cell(row=row_num, column=col_num)
            cell.value = cell_value

    book.save(response)
    return response


def topology_view(request):
    form = SelectTopologyForm()

    topology = request.POST.get('topology')
    # qs_topology = Topology.objects.filter(pk=topology)
    # print(qs_topology)

    # и begin и end по дезайну принад
    edges = [{"from": edge.begin.pk,
              "to": edge.end.pk,
              "label": str(edge.cost),
              "font": {"aligin": "middle"},
              "title": f'pk: {edge.id}, meta: {edge.meta_data}'} for edge in Edge.objects.select_related().filter(begin__topology=topology)]

    nodes = [{"id": node.pk,
              "label": node.label,
              "x": node.x * 50,
              "y": node.y * 50,
              "title": str({**{'pk': node.pk}, **node.meta_data})} for node in Node.objects.select_related().filter(topology=topology)]
    for i in range(len(nodes)):
        if 'DR' in nodes[i]['label']:
            nodes[i].update({'shape': 'hexagon', 'size': 10, 'color': {'background': 'pink', 'border': 'purple'}})



    context = {'edges': edges, 'nodes': nodes, 'form': form}
    return render(request, 'topology/topology_view.html', context=context)


class UserLoginView(LoginView):
    template_name = "topology/login.html"
    form_class = AuthUserForm
    # это не работает
    # success_url = reverse_lazy('login')
    # работает код ниже или ещё можно в settings.py задать константу LOGIN_REDIRECT_URL

    def get_success_url(self):
        return reverse_lazy('home')


class UserCreateView(CreateView):
    form_class = RegisterUserForm
    template_name = "topology/register.html"
    success_url = reverse_lazy('login')
    success_msg = "Пользователь успешно создан"

    def form_valid(self, form):
        # сохранить пользователя в DB
        user = form.save()
        login(self.request, user)
        return redirect('home')

def import_topology(request):
    if request.method == 'POST':
        form = ImportTopologyForm(request.POST)
        if form.is_valid():

            # topology = form.cleaned_data['topology']
            topology = Topology(description=datetime.datetime.now())
            topology.save()

            edges_raw = get_lsdb(form.cleaned_data['raw_data'])
            # G = nx.Graph()
            # G.add_edges_from([edge[1:3] for edge in edges_raw])
            # forceatlas2 = ForceAtlas2(
            #     # Behavior alternatives
            #     outboundAttractionDistribution=True,  # Dissuade hubs
            #     linLogMode=False,  # NOT IMPLEMENTED
            #     adjustSizes=False,  # Prevent overlap (NOT IMPLEMENTED)
            #     edgeWeightInfluence=2.0,
            #
            #     # Performance
            #     jitterTolerance=1.0,  # Tolerance
            #     barnesHutOptimize=True,
            #     barnesHutTheta=1.2,
            #     multiThreaded=False,  # NOT IMPLEMENTED
            #
            #     # Tuning
            #     scalingRatio=2.0,
            #     strongGravityMode=False,
            #     gravity=15.0,
            #
            #     # Log
            #     verbose=True)
            # pos = forceatlas2.forceatlas2_networkx_layout(G, pos=None, iterations=200000)

            x = list({edge[1] for edge in edges_raw} | {edge[2] for edge in edges_raw})
            nodes_raw = {node: Node(topology=topology,
                                    label=node,
                                    meta_data={'id': node},
                                    x=randrange(1, 101),
                                    y=randrange(1, 101)) for node in x}
            for node in nodes_raw:
                nodes_raw[node].save()
            edges = [Edge(begin=nodes_raw[edge[1]],
                          end=nodes_raw[edge[2]],
                          cost=int(edge[3]),
                          meta_data=edge[4]) for edge in edges_raw]
            for i in edges:
                i.save()
            calc_layout.delay(topology.pk)
    else:
        form = ImportTopologyForm()
    return render(request, 'topology/topology_import.html', {'form': form})


def nodes_update(request):
    if request.method == 'POST':
        form = UpdateNodesForm(request.POST)
        if form.is_valid():
            x = form.cleaned_data['raw_data'].split('\n')
            data = [i.split(',') for i in x]
            topology = form.cleaned_data['topology']
            print(topology, type(topology))
            for i in data:
                node = Node.objects.get(topology=topology, label=i[0])
                node.label = i[1]
                node.meta_data.update(json.loads(i[2]))
                node.save()
            print(form.cleaned_data['raw_data'])
    else:
        form = UpdateNodesForm()
    return render(request, 'topology/nodes_update.html', {'form': form})


class EdgeDetailedView(DetailView):
    model = Edge


class EdgeCreateView(CreateView):
    model = Edge
    fields = ['begin', 'end', 'cost', 'meta_data']


class EdgeUpdateView(UpdateView):
    model = Edge
    fields = ['begin', 'end', 'cost', 'meta_data']


class EdgeDeleteView(DeleteView):
    model = Edge
    success_url = "/edge/list"


class EdgeListView(ListView):
    #model = Edge
    form = SelectTopologyForm()
    extra_context = {'form': form}

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)

    def get_queryset(self):
        topology = self.request.POST.get('topology')
        if topology:
            return Edge.objects.select_related().filter(begin__topology__pk__exact=topology)
        # return Edge.objects.select_related().all()
        return Edge.objects.none()


class NodeCreateView(CreateView):
    model = Node
    fields = ['label', 'meta_data', 'topology']


class NodeUpdateView(UpdateView):
    model = Node
    fields = ['label', 'meta_data', 'topology']


class NodeDeleteView(DeleteView):
    model = Node
    success_url = "/node/list"


class NodeDetailedView(DetailView):
    model = Node


class NodeListView(ListView):
    form = SelectTopologyForm()
    extra_context = {'form': form}

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)

    def get_queryset(self):
        topology = self.request.POST.get('topology')
        if topology:
            return Node.objects.select_related().filter(topology__pk__exact=topology)
        # return Node.objects.select_related().all()
        return Node.objects.none()


class TopologyCreateView(CreateView):
    model = Topology
    fields = ['description']


class TopologyUpdateView(UpdateView):
    model = Topology
    fields = ['description', 'is_processed']


class TopologyDeleteView(DeleteView):
    model = Topology
    success_url = "/topology/list"


class TopologyDetailedView(DetailView):
    model = Topology


class TopologyListView(ListView):
    model = Topology


def run(request, pk):
    calc_layout.delay(pk)
    return HttpResponseRedirect(reverse('topology_list'))

